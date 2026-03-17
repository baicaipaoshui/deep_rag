from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    import fitz
except Exception:  # pragma: no cover - optional dependency
    fitz = None

try:
    import openpyxl
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None

try:
    import pandas as pd
except Exception:  # pragma: no cover - optional dependency
    pd = None

from mcp_server.core.index_manager import IndexManager


class FileReader:
    def __init__(self, index_manager: IndexManager) -> None:
        self.index_manager = index_manager

    def read_section(
        self,
        file_name: str,
        section_id: str,
        max_chars: int = 2000,
        filter_values: dict[str, list[str]] | None = None,
        max_rows: int | None = None,
    ) -> dict[str, Any]:
        section = self.index_manager.get_section(file_name, section_id)
        if section is None:
            return {
                "error": "not_found",
                "message": f"section not found: {file_name}:{section_id}",
                "detail": None,
            }

        file_entry = self.index_manager.get_file_entry(file_name)
        if not file_entry:
            return {
                "error": "not_found",
                "message": f"file not found in navigation index: {file_name}",
                "detail": None,
            }
        file_path = self.index_manager.get_absolute_path(file_name)
        if file_path is None or not file_path.exists():
            return {
                "error": "not_found",
                "message": f"source file not found: {file_name}",
                "detail": str(file_path) if file_path else None,
            }

        file_format = file_entry.get("file_format", "")
        if file_format == "markdown":
            content, tables = self._read_markdown(file_path, section)
        elif file_format == "pdf":
            content, tables = self._read_pdf(file_path, section)
        else:
            content, tables = self._read_excel(file_path, section, filter_values, max_rows)

        content = content or ""
        truncated = len(content) > max_chars
        output_content = content[:max_chars] if truncated else content
        return {
            "file_name": file_name,
            "section_id": section_id,
            "heading": section.get("heading", ""),
            "content": output_content,
            "tables": tables,
            "char_count": len(output_content),
            "truncated": truncated,
        }

    @staticmethod
    def _read_markdown(path: Path, section: dict[str, Any]) -> tuple[str, list[dict[str, Any]]]:
        line_start = int(section.get("line_start", 1))
        line_end = int(section.get("line_end", line_start))
        lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
        block = "\n".join(lines[max(0, line_start - 1) : max(line_start, line_end)])
        return block.strip(), []

    @staticmethod
    def _read_pdf(path: Path, section: dict[str, Any]) -> tuple[str, list[dict[str, Any]]]:
        pages_str = str(section.get("pages") or section.get("metadata", {}).get("page") or "1")
        try:
            page_num = int(str(pages_str).split("-")[0])
        except ValueError:
            page_num = 1

        if fitz is None:
            return f"[PDF parsing unavailable] {path.name} page {page_num}", []
        with fitz.open(path) as doc:
            if page_num < 1 or page_num > len(doc):
                return "", []
            page = doc[page_num - 1]
            return (page.get_text("text") or "").strip(), []

    @staticmethod
    def _read_excel(
        path: Path,
        section: dict[str, Any],
        filter_values: dict[str, list[str]] | None,
        max_rows: int | None,
    ) -> tuple[str, list[dict[str, Any]]]:
        sheet_name = str(section.get("sheet_name") or section.get("heading") or "")
        metadata = section.get("metadata", {}) or {}
        section_type = str(metadata.get("section_type", "summary")).strip().lower()
        if openpyxl is None:
            return f"[Excel parsing unavailable] {path.name}:{sheet_name}", []

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

        if not rows:
            return "", []
        headers = FileReader._normalize_excel_cells(rows[0])
        headers = [col if col else f"col_{idx + 1}" for idx, col in enumerate(headers)]
        data_rows = rows[1:]
        normalized_rows = [
            FileReader._normalize_excel_cells(row)
            for row in data_rows
            if any(v is not None and str(v).strip() for v in row)
        ]

        if filter_values and pd is not None and headers and normalized_rows:
            width = len(normalized_rows[0])
            df = pd.DataFrame(normalized_rows, columns=headers[:width])
            for col, values in filter_values.items():
                if col in df.columns and values:
                    df = df[df[col].astype(str).isin([str(v) for v in values])]
            normalized_rows = df.astype(str).values.tolist()

        if section_type == "row_window":
            start_idx = max(0, int(metadata.get("row_start", 2)) - 2)
            end_row = int(metadata.get("row_end", start_idx + 1))
            end_idx = max(start_idx, end_row - 1)
            normalized_rows = normalized_rows[start_idx : end_idx + 1]
        elif section_type == "column_chunk":
            start_col = max(1, int(metadata.get("column_start", 1)))
            end_col = max(start_col, int(metadata.get("column_end", start_col)))
            start_idx = start_col - 1
            end_idx = end_col
            headers = headers[start_idx:end_idx]
            normalized_rows = [row[start_idx:end_idx] for row in normalized_rows]
        elif section_type == "summary":
            normalized_rows = normalized_rows[:20]

        if max_rows is not None:
            normalized_rows = normalized_rows[:max_rows]

        lines = []
        if headers:
            lines.append(" | ".join(headers))
        for row in normalized_rows:
            lines.append(" | ".join(row[: len(headers)]))

        table = {
            "title": sheet_name or "Sheet1",
            "headers": headers,
            "rows": normalized_rows[:30],
        }
        return "\n".join(lines).strip(), [table]

    @staticmethod
    def _normalize_excel_cells(values: list[Any] | tuple[Any, ...]) -> list[str]:
        return ["" if cell is None else str(cell).strip() for cell in values]
